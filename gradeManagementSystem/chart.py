from second import calculator
import root
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

calculator.fun1()
calculator.pass_fail()
calculator.print_max_average()
#print(calculator.mark)

wb=xl.load_workbook('StudentInfo.xlsx')
sheet=wb['Sheet1']
row=2
column=2
for name1 in calculator.student_with_max_avg:
        cell = sheet.cell(row, 1)
        cell.value = name1
        row = row + 1


row=2
column=2
for grade1 in calculator.mark:
        for grade2 in grade1:
                        cell = sheet.cell(row, column)
                        cell.value = grade2
                        column = column + 1
        column=2
        row = row + 1

row=2
for avg1 in calculator.avg_list:
    cell = sheet.cell(row,6)
    cell.value = avg1
    cell = sheet.cell(row, 7)
    if avg1>=60:
        cell.value='Passed'
    else:
        cell.value='Failed'
    row = row + 1


values=Reference(sheet,
                 min_row=2,
                 max_row=len(calculator.avg_list)+1,
                 min_col=6,
                 max_col=7)
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'H1')

wb.save('StudentInfo2.xlsx')